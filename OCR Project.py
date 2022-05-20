#System variable and io handling
import os
import csv

from shutil import copyfile, rmtree
from re import match
import sys
import random


#import configparser
#Regular expression handlings
import multiprocessing
from multiprocessing import Process , Queue, Manager
import queue 
import subprocess
#Get timestamp

from datetime import datetime
from tkinter.constants import COMMAND
#function difination

from urllib.parse import urlparse

#GUI
from tkinter.ttk import Entry, Label, Treeview, Scrollbar, OptionMenu
from tkinter.ttk import Button, Notebook, Radiobutton
from tkinter.ttk import Progressbar, Style

from tkinter import Tk, Frame
from tkinter import Menu, filedialog, messagebox
from tkinter import Text, colorchooser
from tkinter import IntVar, StringVar
from tkinter import W, E, S, N, END, RIGHT, HORIZONTAL, NO, CENTER
from tkinter import WORD, NORMAL, BOTTOM, X, TOP, BOTH, Y
from tkinter import DISABLED

from tkinter import scrolledtext 

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
from openpyxl.styles import Color, PatternFill, Font

from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string

from openpyxl.drawing.image import Image

import webbrowser

from libs.configmanager import ConfigLoader
from libs.version import get_version
from libs.tkinter_extension import AutocompleteCombobox

import cv2
import pytesseract

#import Levenshtein as lev
from rapidfuzz.distance import Levenshtein as lev_distance
from rapidfuzz.distance import Indel as lev_ratio

#from document_toolkit_function.py import *

DELAY1 = 20

ToolDisplayName = "OCR Project"
tool_name = 'ocr'
rev = 1104
a,b,c,d = list(str(rev))
VerNum = a + '.' + b + '.' + c + chr(int(d)+97)

version = ToolDisplayName  + " " +  VerNum 

#**********************************************************************************
# UI handle ***********************************************************************
#**********************************************************************************

class OCR_Project(Frame):
	def __init__(self, Root, Queue = None, Manager = None,):
		
		Frame.__init__(self, Root) 
		#super().__init__()
		self.parent = Root 
		self.parent.protocol("WM_DELETE_WINDOW", self.on_closing)
		# Queue
		self.Process_Queue = Queue['Process_Queue']
		self.Result_Queue = Queue['Result_Queue']
		self.Status_Queue = Queue['Status_Queue']
		self.Debug_Queue = Queue['Debug_Queue']
		self.Manager = Manager['Default_Manager']

		self.Options = {}

		# XLSX Optmizer
		self.Optimize_Folder = ""
		self.Optimize_FileList = ""
		# XLSX Comparision
		self.Compare_Folder_Old = ""
		self.Compare_File_List_Old = ""
		self.Compare_Folder_New = ""
		self.Compare_File_List_New = ""

		# UI Variable
		self.Button_Width_Full = 20
		self.Button_Width_Half = 15
		
		self.PadX_Half = 5
		self.PadX_Full = 10
		self.PadY_Half = 5
		self.PadY_Full = 10
		self.StatusLength = 120
		self.AppLanguage = 'en'

		self.OCR_File_Path = None

		self.Path_Size = 60

		self.init_App_Setting()
		
		self.App_LanguagePack = {}
		

		if self.AppLanguage != 'kr':
			from libs.languagepack import LanguagePackEN as LanguagePack
		else:
			from libs.languagepack import LanguagePackKR as LanguagePack

		self.LanguagePack = LanguagePack

		# Init function

		self.parent.resizable(False, False)
		self.parent.title(version)
		# Creating Menubar 
		
		#**************New row#**************#
		self.Notice = StringVar()
		self.Debug = StringVar()
		self.Progress = StringVar()
			
		#Generate UI
		self.Generate_Menu_UI()
		self.Generate_Tab_UI()
		self.init_UI()
		
		self.init_UI_Data()

	def on_closing(self):
		if messagebox.askokcancel("Quit", "Do you want to quit?"):
			self.parent.destroy()
			try:
				self.OCR_Scan_Process.terminate()
			except:
				pass	

	# UI init
	def init_UI(self):
	
		self.Generate_OCR_Tool_UI(self.OCR_TOOL)

		self.Generate_OCR_Setting_UI(self.OCR_SETTING)



	def Generate_Menu_UI(self):
		menubar = Menu(self.parent) 
		# Adding File Menu and commands 
		'''
		file = Menu(menubar, tearoff = 0)
		
		# Adding Load Menu 
		menubar.add_cascade(label =  self.LanguagePack.Menu['File'], menu = file) 
		file.add_command(label =  self.LanguagePack.Menu['LoadTM'], command = self.Menu_Function_Select_TM) 
		file.add_separator() 
		file.add_command(label =  self.LanguagePack.Menu['CreateTM'], command = self.Menu_Function_Create_TM)
		file.add_separator() 
		file.add_command(label =  self.LanguagePack.Menu['Exit'], command = self.parent.destroy) 
		'''
		# Adding Help Menu
		help_ = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Help'], menu = help_) 
		help_.add_command(label =  self.LanguagePack.Menu['GuideLine'], command = self.Menu_Function_Open_Main_Guideline) 
		help_.add_separator()
		help_.add_command(label =  self.LanguagePack.Menu['About'], command = self.Menu_Function_About) 
		self.parent.config(menu = menubar)

		# Adding Help Menu
		language = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Language'], menu = language) 
		language.add_command(label =  self.LanguagePack.Menu['Hangul'], command = self.SetLanguageKorean) 
		language.add_command(label =  self.LanguagePack.Menu['English'], command = self.SetLanguageEnglish) 
		self.parent.config(menu = menubar) 	

	def Generate_Tab_UI(self):
		
		self.TAB_CONTROL = Notebook(self.parent)

		self.OCR_TOOL = Frame(self.TAB_CONTROL)
		#self.TAB_CONTROL.add(self.AutoTest, text= self.LanguagePack.Tab['AutomationTest'])
		self.TAB_CONTROL.add(self.OCR_TOOL, text= 'OCR Project')
		
		
		self.OCR_SETTING = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.OCR_SETTING, text= 'OCR Setting')

		self.TAB_CONTROL.pack(expand=1, fill="both")
		return

	#STABLE
	def Generate_OCR_Tool_UI(self, Tab):
		'''
		Create main tab
		'''
		
		Row=1
		self.Str_OCR_Image_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['ImageSource']).grid(row=Row, column=1, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = self.Path_Size, state="readonly", textvariable=self.Str_OCR_Image_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=2, columnspan=7, padx=0, pady=5, sticky=E+W)
		 
		Btn_Browse_Image = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_OCR_Browse_Image_Data)
		Btn_Browse_Image.grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)

		Row+=1
		self.Str_OCR_Config_Path = StringVar()
		Label(Tab, text= self.LanguagePack.Label['ScanConfig']).grid(row=Row, column=1, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = self.Path_Size, state="readonly", textvariable=self.Str_OCR_Config_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=2, columnspan=7, padx=0, pady=5, sticky=E+W)
		
		Btn_Browse_Setting = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_OCR_Browse_Config_File)
		Btn_Browse_Setting.grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row+=1
		Label(Tab, width= 10 ,text= self.LanguagePack.Label['CenterX']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_CenterX = Text(Tab, width = 10, height=1) #
		self.Str_CenterX.grid(row=Row, column=2, padx=0, pady=5, sticky=W)

		self.Str_CenterX.bind("<Tab>", self.entry_next)	

		Label(Tab, width= 10 , text= self.LanguagePack.Label['CenterY']).grid(row=Row, column=3, padx=0, pady=5, sticky=W)
		self.Str_CenterY = Text(Tab, width=10, height=1) #
		self.Str_CenterY.grid(row=Row, column=4, padx=0, pady=5, sticky=W)
		self.Str_CenterY.bind("<Tab>", self.entry_next)	

		self.Region_Type = IntVar()
		self.Option_Text_Update = Radiobutton(Tab, width= 10, text= "Main Area", value=1, variable=self.Region_Type, command=None)
		self.Option_Text_Update.grid(row=Row, column=5,columnspan=2,padx=0, pady=5, sticky=W)
		self.Region_Type.set(1)

		Btn_Input_Area = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['AddAreaWithText'], command= self.Btn_OCR_Input_Text_Area)
		Btn_Input_Area.grid(row=Row, column=7, padx=0, pady=5, sticky=W)

		Label(Tab, text= self.LanguagePack.Label['BrowseType']).grid(row=Row, column=8, rowspan=2, pady=5, sticky=W)
		Radiobutton(Tab, width= 10, text=  self.LanguagePack.Option['Folder'], value=1, variable=self.Browse_Type, command=self.OCR_Setting_Set_Browse_Type).grid(row=Row, column=9,columnspan=2,padx=0, pady=5, sticky=E)
	
		Row+=1
		Label(Tab, width= 10 , text= self.LanguagePack.Label['Height']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_Height = Text(Tab, width=10, height=1) #
		self.Str_Height.grid(row=Row, column=2, padx=0, pady=5, sticky=W)
		self.Str_Height.bind("<Tab>", self.entry_next)	
	
		Label(Tab, width= 10 , text= self.LanguagePack.Label['Weight']).grid(row=Row, column=3, padx=0, pady=5, sticky=W)
		self.Str_Weight = Text(Tab, width = 10, height=1) #
		self.Str_Weight.grid(row=Row, column=4, padx=0, pady=5, sticky=W)
		self.Str_Weight.bind("<Tab>", self.entry_next)	

		# Update Template/ Text gacha
		self.Option_Image_Update = Radiobutton(Tab, width= 10, text= "Image Area", value=2, variable=self.Region_Type, command=None)
		self.Option_Image_Update.grid(row=Row, column=5,columnspan=2,padx=0, pady=5, sticky=W)
		#self.Option_Image_Update.configure(state=DISABLED)

		self.Btn_Update_Area = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['SaveConfig'], command= self.Btn_OCR_Update_Area)
		self.Btn_Update_Area.grid(row=Row, column=7, padx=0, pady=5, sticky=W)
		self.Btn_Update_Area.configure(state=DISABLED)

		Radiobutton(Tab, width= 10, text= self.LanguagePack.Option['File'], value=2, variable=self.Browse_Type, command=self.OCR_Setting_Set_Browse_Type).grid(row=Row, column=9,columnspan=2, padx=0, pady=5, sticky=E)
		
		
		Row+=1
		#
		self.Generate_Treeview_Advanced_UI(Tab, Row)

		Btn_Select_Area = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['SelectArea'], command= self.Btn_OCR_Select_Area_Advanced)
		Btn_Select_Area.grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=W)
		
		Row+=1
		Btn_Preview_Area = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['PreviewArea'], command= self.Btn_OCR_Preview_Areas)
		Btn_Preview_Area.grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row+=1
		Btn_Save_Setting = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SaveConfig'], command= self.Btn_OCR_Save_Config_File)
		Btn_Save_Setting.grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=W)

		Row+=1
		self.Btn_Open_Result = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['OpenOutput'], command= self.Open_OCR_Result_Folder)
		self.Btn_Open_Result.grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		self.Btn_Open_Result.configure(state=DISABLED)

		Row+=1
		Btn_Update_Language = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['UpdateLanguage'], command= self.Btn_OCR_Update_Working_Language)
		Btn_Update_Language.grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=W)
		#Btn_Execute.configure(state=DISABLED)

		Row+=1
		Label(Tab, text= self.LanguagePack.Label['Debug']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Debugger = scrolledtext.ScrolledText(Tab, width=110, height=5, undo=False, wrap=WORD, )
		self.Debugger.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=W+E+N+S)

		Row += 1
		Label(Tab, text= self.LanguagePack.Label['WorkingRes']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		Radiobutton(Tab, width= 10, text=  '720p', value=1, variable=self.Resolution, command= self.OCR_Setting_Set_Working_Resolution).grid(row=Row, column=2, padx=0, pady=5, sticky=W)
		Radiobutton(Tab, width= 10, text=  '1080p', value=2, variable=self.Resolution, command= self.OCR_Setting_Set_Working_Resolution).grid(row=Row, column=3, padx=0, pady=5, sticky=W)
	
		Row += 1

		Label(Tab, text= self.LanguagePack.Label['ScanType']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		_scan_type = ['', 'Text only', 'Image only', 'Image and Text', 'DB Create']
		Option_ScanType = OptionMenu(Tab, self.ScanType, *_scan_type, command = self.OCR_Setting_Set_Scan_Type)
		Option_ScanType.config(width=20)
		Option_ScanType.grid(row=Row, column=2,padx=0, pady=5, sticky=W)

		Label(Tab, text= self.LanguagePack.Label['WorkingLang']).grid(row=Row, column=3, padx=5, pady=5, sticky=W)
		self.option_working_language = AutocompleteCombobox(Tab)
		self.option_working_language.Set_Entry_Width(10)
		self.option_working_language.grid(row=Row, column=4, padx=5, pady=5, sticky=W)
		
		Row+=1
		Label(Tab, text= self.LanguagePack.Label['Progress']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.progressbar = Progressbar(Tab, orient=HORIZONTAL, length=800,  mode='determinate')
		self.progressbar["maximum"] = 1000
		self.progressbar.grid(row=Row, column=2, columnspan=7, padx=5, pady=5, sticky=E+W)

		Btn_Execute = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['Scan'], command= self.Btn_OCR_Execute)
		Btn_Execute.grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=W)

	def Generate_OCR_Setting_UI(self, Tab):
		''''
		Create Setting Tab
		'''
		Row = 1
		Label(Tab, text= self.LanguagePack.Label['TesseractPath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_TesseractPath = Entry(Tab,width = 100, state="readonly", textvariable=self.TesseractPath)
		self.Text_TesseractPath.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=E+W)
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Select_Tesseract_Path).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['TesseractDataPath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_TesseractDataPath = Entry(Tab,width = 100, state="readonly", textvariable=self.TesseractDataPath)
		self.Text_TesseractDataPath.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=E+W)
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Select_Tesseract_Data_Path).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['DBPath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_DB_Path = Entry(Tab,width = 100, state="readonly", textvariable=self.DBPath)
		self.Text_DB_Path.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=E+W)
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Select_DB_Path).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
	def Generate_Treeview_Simple_UI(self, Tab, Row = 5):
		TreeView_Row = 5
		self.Treeview = Treeview(Tab)
		self.Focused_Item = None
		self.Treeview.grid(row=Row, column=1, columnspan=8, rowspan=TreeView_Row, padx=5, pady=5, sticky = N+S+W+E)
		verscrlbar = Scrollbar(Tab, orient ="vertical", command = self.Treeview.yview)
		self.Treeview.configure( yscrollcommand=verscrlbar.set)
	
		self.Treeview.Scrollable = True
		self.Treeview['columns'] = ('X', 'Y', 'W', 'H')

		self.Treeview.column('#0', width=0, stretch=NO)
		self.Treeview.heading('#0', text='', anchor=CENTER)

		for column in self.Treeview['columns']:
			self.Treeview.column(column, anchor=CENTER, width=100)
			self.Treeview.heading(column, text=column, anchor=CENTER)
		

		verscrlbar.grid(row=Row, column=8, rowspan=TreeView_Row,  sticky = N+S+E)
		Tab.grid_columnconfigure(11, weight=0, pad=0)
		styles = Style()
		styles.configure('Treeview',rowheight=15)

		self.Treeview.bind("<Delete>", self.delete_treeview_line)	
		self.Treeview.bind("<Double-1>", self.Treeview_OCR_Select_Row)
	
	def Generate_Treeview_Advanced_UI(self, Tab, Row = 5):
		TreeView_Row = 5
		self.Treeview = Treeview(Tab)
		self.Focused_Item = None
		self.Treeview.grid(row=Row, column=1, columnspan=8, rowspan=TreeView_Row, padx=5, pady=5, sticky = N+S+W+E)
		verscrlbar = Scrollbar(Tab, orient ="vertical", command = self.Treeview.yview)
		self.Treeview.configure( yscrollcommand=verscrlbar.set)
	
		self.Treeview.Scrollable = True
		self.Treeview['columns'] = ('X', 'Y', 'W', 'H', 'X1', 'Y1', 'W1', 'H1')

		self.Treeview.column('#0', width=0, stretch=NO)
		self.Treeview.heading('#0', text='', anchor=CENTER)

		for column in self.Treeview['columns']:
			self.Treeview.column(column, anchor=CENTER, width=100)
			self.Treeview.heading(column, text=column, anchor=CENTER)

		verscrlbar.grid(row=Row, column=8, rowspan=TreeView_Row,  sticky = N+S+E)
		Tab.grid_columnconfigure(11, weight=0, pad=0)
		styles = Style()
		styles.configure('Treeview',rowheight=15)

		self.Treeview.bind("<Delete>", self.delete_treeview_line)	
		self.Treeview.bind("<Double-1>", self.Treeview_OCR_Select_Row)



###########################################################################################
# Treeview FUNCTION
###########################################################################################

	def delete_treeview_line(self, event):
		'''
		Function activate when select an entry from a Treeview and press Delete btn
		'''
		selected = self.Treeview.selection()
		to_remove = []
		for child_obj in selected:
			child = self.Treeview.item(child_obj)
			tm_index = child['values'][0]
			to_remove.append(tm_index)
			self.Treeview.delete(child_obj)

	# Obsoleted.
	def double_right_click_treeview(self, event):
		'''
		Function activate when double click an entry from Treeview
		'''
		focused = self.Treeview.focus()
		child = self.Treeview.item(focused)
		self.Debugger.insert("end", "\n")
		self.Debugger.insert("end", 'Korean: ' + str(child["text"]))
		self.Debugger.insert("end", "\n")
		self.Debugger.insert("end", 'English: ' + str(child["values"][0]))
		self.Debugger.yview(END)


	# Nam will check
	def load_tm_list(self):
		"""
		When clicking the [Load] button in TM Manager tab
		Display the pair languages in the text box.
		"""
		self.remove_treeview()
		
		_area_list = []

		for location in _area_list:	
			try:
				self.Treeview.insert('', 'end', text= '', values=( str(location['index']), str(location['x']), str(location['y']), str(location['h']), str(location['w'])))
			except:
				pass

	def add_treeview_row(self, location):
		'''
		Add a row to the current Treeview
		'''
		self.Treeview.insert('', 'end', text= '', values=(str(location[0]), str(location[1]), str(location[2]), str(location[3])))

	def remove_treeview(self):
		for i in self.Treeview.get_children():
			self.Treeview.delete(i)

###########################################################################################
# MENU FUNCTION
###########################################################################################

	def Menu_Function_About(self):
		messagebox.showinfo("About....", "Designer: Mr. 박찬혁\r\nDeveloper: Evan")

	def Show_Error_Message(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def SaveAppLanguage(self, language):
		self.Write_Debug(self.LanguagePack.ToolTips['AppLanuageUpdate'] + " "+ language) 
		self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'app_lang', language)

	def SetLanguageKorean(self):
		self.AppLanguage = 'kr'
		self.SaveAppLanguage(self.AppLanguage)
	
	def SetLanguageEnglish(self):
		self.AppLanguage = 'en'
		self.SaveAppLanguage(self.AppLanguage)

	def Function_Correct_Path(self, path):
		return str(path).replace('/', '\\')
	
	def Menu_Function_Open_Main_Guideline(self):
		webbrowser.open_new(r"https://confluence.nexon.com/display/NWMQA/OCR+%28Optical+Character+Recognition%29+Tool")

	def onExit(self):
		self.quit()

	def init_App_Setting(self):

		self.DB_Path = StringVar()
		self.TesseractPath = StringVar()
		self.TesseractDataPath = StringVar()
		self.WorkingLanguage = StringVar()
		self.language_list = ['']

		self.DBPath = StringVar()

		self.Browse_Type = IntVar()

		self.Resolution = IntVar()
		self.CurrentDataSource = StringVar()

		self.ScanType = StringVar()

		self.Notice = StringVar()

		self.AppConfig = ConfigLoader()
		self.Configuration = self.AppConfig.Config
		self.AppLanguage  = self.Configuration['OCR_TOOL']['app_lang']

		_tesseract_path = self.Configuration['OCR_TOOL']['tess_path']
		pytesseract.pytesseract.tesseract_cmd = str(_tesseract_path)
		self.TesseractPath.set(_tesseract_path)

		_tesseract_data_path = self.Configuration['OCR_TOOL']['tess_data']
		self.TesseractDataPath.set(_tesseract_data_path)

		_db_path = self.Configuration['OCR_TOOL']['db_path']
		self.DBPath.set(_db_path)


		_browse_type = self.Configuration['OCR_TOOL']['browsetype']
		self.Browse_Type.set(_browse_type)

		_resolution = self.Configuration['OCR_TOOL']['resolution']
		self.Resolution.set(_resolution)

		
	def init_UI_Data(self):
		self.Btn_OCR_Update_Working_Language()
		_working_language = self.Configuration['OCR_TOOL']['scan_lang']
		self.option_working_language.set(_working_language)

		_scan_type = self.Configuration['OCR_TOOL']['scan_type']
		self.ScanType.set(_scan_type)
		if _scan_type == 'Image and Text':
			self.Option_Image_Update.configure(state=NORMAL)


	def SaveSetting(self):

		print('Save setting')
		return


###########################################################################################
# General functions
###########################################################################################

	def CorrectPath(self, path):
		if sys.platform.startswith('win'):
			return str(path).replace('/', '\\')
		else:
			return str(path).replace('\\', '//')
	
	def CorrectExt(self, path, ext):
		if path != None and ext != None:
			Outputdir = os.path.dirname(path)
			baseName = os.path.basename(path)
			sourcename = os.path.splitext(baseName)[0]
			newPath = self.CorrectPath(Outputdir + '/'+ sourcename + '.' + ext)
			return newPath

	def Write_Debug(self, text):
		'''
		Function write the text to debugger box and move to the end of the box
		'''
		self.Debugger.insert("end", "\n")
		self.Debugger.insert("end", str(text))

		self.Debugger.yview(END)		

	def entry_next(self, event):
		event.widget.tk_focusNext().focus()
		return("break")



###########################################################################################
# OCR
###########################################################################################
	
	def Btn_OCR_Select_Background_Colour(self):
		colorStr, self.Background_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		if self.Background_Color == None:
			self.Show_Error_Message('Set colour as defalt colour (Yellow)')
			self.Background_Color = 'ffff00'
		else:
			self.Background_Color = self.Background_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return
	
	def Btn_OCR_Select_Font_Colour(self):
		colorStr, self.Font_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Font_Color == None:
			self.Show_Error_Message('Set colour as defalt colour (Yellow)')
			self.Font_Color = 'FF0000'
		else:
			self.Font_Color = self.Font_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_OCR_Browse_Config_File(self):

		_scan_type = self.ScanType.get()

		self.Btn_Open_Result.configure(state=DISABLED)

		config_file = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'], filetypes = (("Config files", "*.csv *.xlsx"), ), multiple = False)	
		
		if os.path.isfile(config_file):
			print('config_file', config_file)
			self.Str_OCR_Config_Path.set(config_file)
			self.remove_treeview()

			all_col = ['x', 'y', 'w', 'h', 'x1', 'y1', 'w1', 'h1']	
			
			with open(config_file, newline='', encoding='utf-8-sig') as csvfile:
				reader = csv.DictReader(csvfile)
				input_location = {}
				for location in reader:
					for col in all_col:
						if col in location:
							input_location[col] = location[col]
						else:
							input_location[col] = 0
					self.Treeview.insert('', 'end', text= '', values=(str(input_location['x']), str(input_location['y']), str(input_location['w']), str(input_location['h']), str(input_location['x1']), str(input_location['y1']), str(input_location['w1']), str(input_location['h1'])))
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	def Btn_OCR_Save_Config_File(self):
		'''
		Save all added scan areas to csv file.
		'''
		_scan_type = self.ScanType.get()

		filename = filedialog.asksaveasfilename(title = "Select file", filetypes = (("Scan Config", "*.csv"),),)
		filename = self.CorrectExt(filename, "csv")
		if filename == "":
			return
		else:
			with open(filename, 'w', newline='') as csvfile:

				fieldnames = ['x', 'y', 'w', 'h', 'x1', 'y1', 'w1', 'h1']
			
				writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
				writer.writeheader()
				for row in self.Treeview.get_children():
					child = self.Treeview.item(row)
					values = child["values"]	
					writer.writerow({'x': values[0], 'y': values[1], 'w': values[2], 'h': values[3], 'x1': values[4], 'y1': values[5], 'w1': values[6], 'h1': values[7]})
			
	def Btn_OCR_Browse_Image_Data(self):
		
		self.Btn_Open_Result.configure(state=DISABLED)
		
		_select_type = self.Browse_Type.get()
		if _select_type == 1:
			self.Btn_OCR_Browse_Image_Folder()
		else:
			self.Btn_OCR_Browse_Image_Files()

	def Btn_OCR_Browse_Image_Folder(self):
			
		folder_name = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'],)	
		if folder_name != "":
			_temp_text_files = os.listdir(folder_name)
			self.OCR_File_Path = []
			for file in _temp_text_files:
				file_path = folder_name + '/' + file
				if os.path.isfile(file_path):
					baseName = os.path.basename(file_path)
					sourcename, ext = os.path.splitext(baseName)
					if ext in ['.jpg','.jpeg','.png']:
						self.OCR_File_Path.append(file_path)

			self.Str_OCR_Image_Path.set(str(folder_name) + '/*')

			self.Write_Debug(self.LanguagePack.ToolTips['DataSelected'] + ": " + str(len(self.OCR_File_Path)))
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	def Btn_OCR_Browse_Image_Files(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'], filetypes = (("Image files", "*.jpg *.jpeg *png"), ), multiple = True)	
		if filename != "":
			self.OCR_File_Path = list(filename)
			self.Str_OCR_Image_Path.set(str(self.OCR_File_Path[0]))
			
			self.Write_Debug(self.LanguagePack.ToolTips['DataSelected'] + ": " + str(len(self.OCR_File_Path)))
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		
	# Obsoleted
	def Btn_OCR_Select_Area(self):
		
		if self.OCR_File_Path != None:
			_index = random.randint(0, len(self.OCR_File_Path)-1)
			if os.path.isfile(self.OCR_File_Path[_index]):
				im = cv2.imread(self.OCR_File_Path[_index])
				
				(_h, _w) = im.shape[:2]
				ratio = 720 / _h
				if ratio != 1:
					width = int(im.shape[1] * ratio)
					height = int(im.shape[0] * ratio)
					dim = (width, height)
					im = cv2.resize(im, dim, interpolation = cv2.INTER_AREA)

				for row in self.Treeview.get_children():
					child = self.Treeview.item(row)
					im = cv2.rectangle(im, (child["values"][0], child["values"][1]), (child["values"][0] + child["values"][2], child["values"][1] + child["values"][3]), (255,0,0), 2)

				location = cv2.selectROI("Sekect scan area", im, showCrosshair=False,fromCenter=False)
				cv2.destroyAllWindows() 
				self.Treeview.insert('', 'end', text= '', values=(str(location[0]), str(location[1]), str(location[2]), str(location[3])))
			else:
				self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])		
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])	

	def Btn_OCR_Preview_Areas(self):

		_scan_type = self.ScanType.get()

		if self.OCR_File_Path != None:
			_index = random.randint(0, len(self.OCR_File_Path)-1)
			if os.path.isfile(self.OCR_File_Path[_index]):
				#im = self.Function_Load_Img(self.OCR_File_Path[_index])
				im = cv2.imread(self.OCR_File_Path[_index])
				im, ratio = self.Resize_Image_by_ratio(im)

				for row in self.Treeview.get_children():
					child = self.Treeview.item(row)
					areas = child["values"]

					for area in areas:
						area = area * ratio

					if _scan_type in ['Image and Text', 'DB Create', 'Image only']:
						if areas[6] > 0 and areas[7] >0:
							im = cv2.rectangle(im, (areas[4], areas[5]), (areas[4] + areas[6], areas[5] + areas[7]), (255,255,0), 2)
					if _scan_type in ['Image and Text', 'DB Create', 'Text only']:
						if areas[2] > 0 and areas[3] >0:	
							im = cv2.rectangle(im, (areas[0], areas[1]), (areas[0] + areas[2], areas[1] + areas[3]), (255,0,0), 2)
					
				cv2.imshow("Display ratio: " + str(int(ratio*100)) + "%", im)
				cv2.waitKey(0)
			else:
				self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])		
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])	


	def Btn_OCR_Select_Area_Advanced(self):
		_scan_type = self.ScanType.get()

		if self.OCR_File_Path != None:
			_index = random.randint(0, len(self.OCR_File_Path)-1)
			if os.path.isfile(self.OCR_File_Path[_index]):
				im = cv2.imread(self.OCR_File_Path[_index])
				im, ratio = self.Resize_Image_by_ratio(im)
				'''
				(_h, _w) = im.shape[:2]
				ratio = 720 / _h
				'''
				
				if ratio != 1:
					width = int(im.shape[1] * ratio)
					height = int(im.shape[0] * ratio)
					dim = (width, height)
					im = cv2.resize(im, dim, interpolation = cv2.INTER_AREA)

				for row in self.Treeview.get_children():
					child = self.Treeview.item(row)
					areas = child["values"]
					print('1', areas)
					for area in areas:
						area = area * ratio
					print('2', areas)
					if _scan_type in ['Image and Text', 'DB Create', 'Text only']:
						im = cv2.rectangle(im, (areas[0], areas[1]), (areas[0] + areas[2], areas[1] + areas[3]), (255,0,0), 2)
					if _scan_type in ['Image and Text', 'DB Create', 'Image only']:
						im = cv2.rectangle(im, (areas[4], areas[5]), (areas[4] + areas[6], areas[5] + areas[7]), (255,255,0), 2)
				location = [0,0,0,0]
				location2 = [0,0,0,0]
				if _scan_type in ['Image and Text', 'DB Create', 'Text only']:
					location = cv2.selectROI("Select TEXT area", im, showCrosshair=False,fromCenter=False)
					im = cv2.rectangle(im, (location[0], location[1]), (location[0] + location[2], location[1] + location[3]), (255,0,0), 2)
					cv2.destroyAllWindows() 
				if _scan_type in ['Image and Text', 'DB Create', 'Image only']:
					location2 = cv2.selectROI("Select COMPONENT area", im, showCrosshair=False,fromCenter=False)
					cv2.destroyAllWindows() 
				#imCrop = im[int(area[1]/ratio):int(area[1]/ratio + area[3]/ratio), int(area[0]/ratio):int(area[0]/ratio + area[2]/ratio)]
				for area in location:
					area = area/ratio
				for area in location2:
					area = area/ratio
				print(location)
				print(location2)
				self.Treeview.insert('', 'end', text= '', values=(str(location[0]), str(location[1]), str(location[2]), str(location[3]), str(location2[0]), str(location2[1]), str(location2[2]), str(location2[3]) ))
				
			else:
				self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])		
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])				

	def Resize_Image_by_ratio(self, img):
		global WIDTH, HEIGHT
		
		(_h, _w) = img.shape[:2]
		_ratio = 1
		while True:
			temp_w = _w * _ratio
			temp_h = _h * _ratio
			if temp_h > HEIGHT*0.75 or temp_w > WIDTH*0.75:
				_ratio = int(_ratio * 90)/100
			else:
				break
		
		if _ratio != 1:
			width = int(img.shape[1] * _ratio)
			height = int(img.shape[0] * _ratio)
			dim = (width, height)
			img = cv2.resize(img, dim, interpolation = cv2.INTER_AREA)
		
		actual_ratio = _ratio

		return img, actual_ratio	

	def Btn_OCR_Input_Text_Area(self):

		self.Focused_Item = self.Treeview.focus()
		child = self.Treeview.item(self.Focused_Item)
		self.Btn_Update_Area.configure(state=NORMAL)

		_x = self.Str_CenterX.get("1.0", END).replace('\n', '')
		if _x == '': _x = 0
		_y = self.Str_CenterY.get("1.0", END).replace('\n', '')
		if _y == '': _y = 0
		_w = self.Str_Weight.get("1.0", END).replace('\n', '')
		if _w == '': _w = 0
		_h = self.Str_Height.get("1.0", END).replace('\n', '')
		if _h == '': _h = 0
		
		if self.Region_Type.get() == 1:
			self.Treeview.insert('', 'end', text= '', values=(str(int(_x)), str(int(_y)), str(int(_w)), str(int(_h)), child["values"][4], child["values"][5], child["values"][6], child["values"][7]))
		else:
			self.Treeview.insert('', 'end', text= '', values=(child["values"][0], child["values"][1], child["values"][2], child["values"][3], str(int(_x)), str(int(_y)), str(int(_w)), str(int(_h))))

		#self.Update_Treeview_Advanced_UI()

	def Treeview_OCR_Select_Row(self, event):
		'''
		Function activate when double click an entry from Treeview
		'''
		self.Focused_Item = self.Treeview.focus()
		child = self.Treeview.item(self.Focused_Item)
		self.Btn_Update_Area.configure(state=NORMAL)

		
		if self.Region_Type.get() == 1:
			x = 0
		else:
			x = 4	
		self.Str_CenterX.delete("1.0", END)
		try:
			self.Str_CenterX.insert("end", child["values"][x])
		except:
			self.Str_CenterX.insert("end", 0)
			
		self.Str_CenterY.delete("1.0", END)
		try:
			self.Str_CenterY.insert("end", child["values"][x+1])
		except:
			self.Str_CenterY.insert("end", 0)

		self.Str_Weight.delete("1.0", END)
		try:
			self.Str_Weight.insert("end", child["values"][x+2])
		except:
			self.Str_Weight.insert("end", 0)

		self.Str_Height.delete("1.0", END)
		try:
			self.Str_Height.insert("end", child["values"][x+3])
		except:
			self.Str_Height.insert("end", 0)


	def Btn_OCR_Update_Area(self):

		if self.Focused_Item != None:
			child = self.Treeview.item(self.Focused_Item)
			_x = self.Str_CenterX.get("1.0", END).replace('\n', '')
			if _x == '': _x = 0
			_y = self.Str_CenterY.get("1.0", END).replace('\n', '')
			if _y == '': _y = 0
			_w = self.Str_Weight.get("1.0", END).replace('\n', '')
			if _w == '': _w = 0
			_h = self.Str_Height.get("1.0", END).replace('\n', '')
			if _h == '': _h = 0
			
			if self.Region_Type.get() == 1:
				self.Treeview.item(self.Focused_Item, text="", values=(str(int(_x)), str(int(_y)), str(int(_w)), str(int(_h)), child["values"][4], child["values"][5], child["values"][6], child["values"][7]))
			else:
				self.Treeview.item(self.Focused_Item, text="", values=(child["values"][0], child["values"][1], child["values"][2], child["values"][3], str(int(_x)), str(int(_y)), str(int(_w)), str(int(_h))))


			#self.Treeview.item(self.Focused_Item, text="", values=(child["values"]))
			self.Focused_Item = None
			self.Btn_Update_Area.configure(state=DISABLED)
		
	def Function_Load_Img(self, path):
		img = cv2.imread(path)
		(_h, _w) = img.shape[:2]
		_working_res = self.Resolution.get()
		if _working_res == 1:
			_ratio = 720
		else:
			_ratio = 1080

		ratio =  _ratio / _h
		if ratio != 1:
			width = int(img.shape[1] * ratio)
			height = int(img.shape[0] * ratio)
			dim = (width, height)
			img = cv2.resize(img, dim, interpolation = cv2.INTER_AREA)
		
		return img

	

	def Btn_OCR_Update_Working_Language(self):
		_data_ = str(self.TesseractDataPath.get())
		_exe_ = str(self.TesseractPath.get())
		_tessdata_dir_config = '--tessdata-dir ' + "\"" + _data_ + "\""
		pytesseract.pytesseract.tesseract_cmd = _exe_
		#self.language_list = pytesseract.get_languages(config=_tessdata_dir_config)
		try:
			self.language_list = pytesseract.get_languages(config=_tessdata_dir_config)
			self.Write_Debug('Supported language list has been updated!')

		except Exception as e:
			self.Write_Debug('Tess path: ' + str(_exe_))
			self.Write_Debug('Data path: ' + str(_data_))
			self.Write_Debug('Error while updating supported language: ' + str(e))
			self.language_list = ['']

		self.option_working_language.set_completion_list(self.language_list)

	
	def Open_OCR_Result_Folder(self):
		try:
			path = self.Function_Correct_Path(self.Output_Result_Folder)
			_cmd = 'explorer ' + "\"" + str(path) + "\""
			
			subprocess.Popen(_cmd)
		except AttributeError:
			self.Show_Error_Message('Please select source folder.')
			return


	def Btn_OCR_Execute(self):
		'''
		Execute main function
		'''
		Image_Files = self.OCR_File_Path
		Image_Folder =  os.path.dirname( self.OCR_File_Path[0])


		timestamp = Function_Get_TimeStamp()
		
		_db_path = self.DBPath.get()
		_scan_type = self.ScanType.get()
		if _scan_type == 'DB Create':
			self.Output_Result_Folder = os.path.dirname(_db_path)
			print(self.Output_Result_Folder)
		else:
			self.Output_Result_Folder = Image_Folder + '/' + 'Scan_Result_' + str(timestamp)
			if not os.path.isdir(self.Output_Result_Folder):
				os.mkdir(self.Output_Result_Folder)
		output_result_file = self.Output_Result_Folder + '/result.csv'
		_ratio = 720	
		_scan_areas = []
		for row in self.Treeview.get_children():
			child = self.Treeview.item(row)
			_scan_areas.append(child['values'])

		_tess_data = self.TesseractDataPath.get()
		_tess_path = self.TesseractPath.get()
	
		_tess_language = self.option_working_language.get()
		self.OCR_Setting_Set_Working_Language(_tess_language)
		#_tess_language = self.WorkingLanguage.get()

		self.Btn_Open_Result.configure(state=NORMAL)
		db_list = []

		self.OCR_Scan_Process = Process(target=Function_Batch_OCR_Execute, args=(self.Result_Queue, self.Status_Queue, self.Process_Queue, _tess_path,_tess_language, _tess_data, Image_Files, output_result_file, _ratio, _scan_areas,_scan_type, _db_path, ))
		
		self.OCR_Scan_Process.start()
		
		self.progressbar["value"] = 0
		self.progressbar.update()

		self.after(DELAY1, self.Wait_For_OCR_Process)

	def Wait_For_OCR_Process(self):
		if (self.OCR_Scan_Process.is_alive()):
			
			try:
				percent = self.Process_Queue.get(0)
				self.progressbar["value"] = percent
				self.progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Write_Debug(Status)
					
			except queue.Empty:
				pass	
			self.after(DELAY1, self.Wait_For_OCR_Process)
		else:
			while True:
				try:
					percent = self.Process_Queue.get(0)
					self.progressbar["value"] = percent
					self.progressbar.update()
				except queue.Empty:
					break
			while True:
				try:
					Status = self.Status_Queue.get(0)
					if Status != None:	
						self.Write_Debug(Status)
						#print(Status)
				except queue.Empty:
					break
			self.OCR_Scan_Process.terminate()
			self.Write_Debug(self.LanguagePack.ToolTips['Completed'])

###########################################################################################
# OCR Setting
###########################################################################################

	def Btn_Select_Tesseract_Path(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectDB'],filetypes = (("Executable files","*.exe" ), ), )	
		if os.path.isfile(filename):
			_tess_path = self.CorrectPath(filename)
			self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'tess_path', _tess_path, True)
			pytesseract.pytesseract.tesseract_cmd = _tess_path
			self.TesseractPath.set(_tess_path)
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['TessNotSelect'])

	def Btn_Select_Tesseract_Data_Path(self):
		folder_name = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'],)	
		if os.path.isdir(folder_name):
			folder_name = self.CorrectPath(folder_name)
			self.TesseractDataPath.set(folder_name)

			self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'tess_data', folder_name, True)

			self.Write_Debug(self.LanguagePack.ToolTips['DataSelected'] + ": " + folder_name)
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	def Btn_Select_DB_Path(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectDB'],filetypes = (("DB files","*.csv" ), ), )	
		if os.path.isfile(filename):
			_db_path = self.CorrectPath(filename)
			self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'db_path', _db_path, True)
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['TessNotSelect'])

	def OCR_Setting_Set_Scan_Type(self, scan_type):		
		self.ScanType.set(scan_type)
		self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'scan_type', scan_type)
		self.Write_Debug(self.LanguagePack.ToolTips['ScanTypeUpdate'] + str(scan_type) + '.')
		if scan_type == 'DB Create':
			self.Write_Debug(self.LanguagePack.ToolTips['DBCreate'])
		elif scan_type == 'Text only':
			self.Write_Debug(self.LanguagePack.ToolTips['TextScan'])
		elif scan_type == 'Image only':
			self.Write_Debug(self.LanguagePack.ToolTips['ImageScan'])
		elif scan_type == 'Image and Text':
			self.Write_Debug(self.LanguagePack.ToolTips['AdvancedScan'])
	



	def OCR_Setting_Set_Browse_Type(self):
		_browse_type = self.Browse_Type.get()
		if _browse_type == 1:
			_status = 'folder'
		else:
			_status = 'file'
		
		self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'browsetype', _browse_type)

		self.Write_Debug(self.LanguagePack.ToolTips['BrowseTypeUpdate'] + str(_status))

	def OCR_Setting_Set_Working_Resolution(self):
		_resolution_index = self.Resolution.get()
		if _resolution_index == 1:
			self.WorkingResolution = 720
		else:
			self.WorkingResolution = 1080
		
		self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'resolution', _resolution_index)

		self.Write_Debug(self.LanguagePack.ToolTips['SetResolution'] + str(self.WorkingResolution) + 'p')

	def OCR_Setting_Set_Working_Language(self, select_value):
		
		self.AppConfig.Save_Config(self.AppConfig.Ocr_Tool_Config_Path, 'OCR_TOOL', 'scan_lang', select_value)
		
		self.Write_Debug(self.LanguagePack.ToolTips['SetScanLanguage'] + str(select_value))
	
	

###########################################################################################
# Process function - Batch scan
###########################################################################################

def Function_Batch_OCR_Execute(
	Result_Queue, Status_Queue, Process_Queue, tess_path, tess_language, tess_data, image_files, result_file, ratio, scan_areas, scan_type = 'Text only', db_path = [], **kwargs):
	
	advanced_tessdata_dir_config = '--psm 7 --tessdata-dir ' + '"' + tess_data + '"'

	if tess_language == '':
		tess_language = 'kor'

	_output_dir = os.path.dirname(result_file)

	if scan_type == 'Text only':

		_all_image_dir = _output_dir + '\\all_text_images'
		_unique_text_image_dir = _output_dir + '\\unique_text_images'
		current_ratio = 0
		process_ratio = 0.01
		
		initFolder(_all_image_dir)
		initFolder(_unique_text_image_dir)
		
		percent = ShowProgress(process_ratio, 100)
		current_ratio+=process_ratio
		Process_Queue.put(percent)

		Status_Queue.put('Crop image')
		process_ratio = 0.04
		image_info = Function_Crop_All_Text(Process_Queue, image_files, scan_areas, ratio, _all_image_dir, process_ratio, current_ratio)
		current_ratio+=process_ratio


		Status_Queue.put('Filter unique images ('+ str(image_info['count']) + ')')
		process_ratio =0.10
		_draft_result = Function_Filter_Unique_Image(Process_Queue, _all_image_dir, _unique_text_image_dir, process_ratio, current_ratio)
		current_ratio+=process_ratio
		
		count = 0
		for image in _draft_result:
			count = count + _draft_result[image]
	
		result = {}
		
		process_ratio = (1-current_ratio - 0.01)
		_output_dir = os.path.dirname(result_file)
		result_file = _output_dir + '/' + 'Text_Scan_Test_Result' + '.xlsx'
		process_count = 0
		total_process = len(_draft_result.keys())
		Status_Queue.put('Scan text from unique images (' + str(total_process) + ')')
		
		DB  = Function_Import_DB(db_path)
		word_db_list = []
		
		for element in DB:
			word = element[tess_language]
			word_db_list.append(word.replace(' ','').lower())
		
		for image in _draft_result:
			key = str(Function_Get_Text_from_Image(tess_path, tess_language, advanced_tessdata_dir_config, _unique_text_image_dir + '\\' + image))
			_match_type = 'none'
			if len(word_db_list)> 0:
				_temp_text = key.replace(' ','').lower()
				if len(_temp_text) == 0:
					continue
				if _temp_text in word_db_list:
					# exact match
					_index = word_db_list.index(_temp_text)
					_match_type = 'exact'
			
				else:
					# similarity check
					_dist = len(_temp_text)
					_ratio = 0
					_word = ''
					
					for word in word_db_list:
						Distance = lev_distance.distance(_temp_text, word)		
						Ratio = lev_ratio.normalized_similarity(_temp_text, word)
						if Distance <= _dist and Ratio >= _ratio:
							_dist = Distance
							_ratio = Ratio
							_word = word

					if _dist/len(_temp_text) <= 0.2 and _ratio >= 0.8:
						_index = word_db_list.index(_word)
						_key = word_db_list[_index]
						Status_Queue.put('Text has been corrected from: ' + key + ' to ' + _key)
						key = _key
						_match_type = 'corrected'
					elif _dist/len(_temp_text) < 0.34 and _ratio > 0.66 and len(_temp_text) == len(_word):
						_index = word_db_list.index(_word)
						_key = word_db_list[_index]
						Status_Queue.put('Text has been corrected from: ' + key + ' to ' + _key)
						key = _key
						_match_type = 'corrected'	
					
			
			error_count = 0
			if len(key.replace(' ','')) == 0:
				error_count+=1
				key = '[Error_'+ str(error_count) + ']'

			if key in result:
				value = _draft_result[image]
				result[key]['value'] = result[key]['value'] + value
			else:
				value = _draft_result[image]
				result[key] = {}
				result[key]['value'] = value
				result[key]['image'] = image
				result[key]['match_type'] = _match_type

			process_count+=1	
			percent = ShowProgress(process_count, total_process, process_ratio, current_ratio )
			Process_Queue.put(percent)		
		
		current_ratio += process_ratio
		
		Status_Queue.put('Export test result')
		row_height = scan_areas[0][7]
		cell_width = scan_areas[0][2]* (1/6)
		Function_Export_Gacha_Test_Result(result, _unique_text_image_dir,result_file, cell_width, row_height)
		percent = ShowProgress(1, 1)
		Process_Queue.put(percent)

	elif scan_type == 'Image only':	
		
		_all_image_dir = _output_dir + '/all_images'
		_unique_image_dir = _output_dir + '/unique_images'
		result_file = _output_dir + '/' + 'Image_Compare_Test_Result' + '.xlsx'
		DB = Function_Import_DB(db_path)

		_db_dir = os.path.dirname(db_path)
		
		initFolder(_all_image_dir)
		initFolder(_unique_image_dir)

		current_ratio = 0
		process_ratio = 0.4
		process_count = 0
		percent = ShowProgress(process_ratio, 100)
		current_ratio+=process_ratio
		Process_Queue.put(percent)

		Status_Queue.put('Crop image')
		process_ratio = 0.04
		image_info = Function_Crop_All_Image(Process_Queue, image_files, scan_areas, ratio, _all_image_dir, process_ratio, current_ratio)
		current_ratio+=process_ratio


		Status_Queue.put('Filter unique images ('+ str(image_info['count']) + ')')
		process_ratio =0.5
		_draft_result = Function_Filter_Unique_Image(Process_Queue, _all_image_dir, _unique_image_dir, process_ratio, current_ratio)
		current_ratio+=process_ratio
		total_process = len(_draft_result.keys())
		result = []
		#print('_draft_result', _draft_result)
		for unique_image in _draft_result:
			temp_result = {}
			_temp_template = _unique_image_dir + '/' + unique_image
			temp_result['image'] = _temp_template
			temp_result['count'] = _draft_result[unique_image]
			for row in DB:
				_template = _db_dir + '\\' + row['path']
				if os.path.isfile(_template):
					_image_compare_result = Function_Compare_2_Image(_temp_template, _template)
					if _image_compare_result == True:
						_name = row[tess_language]
						temp_result['name'] = _name
						break
					else:
						temp_result['name'] = 'N/A'
				else:
					temp_result['name'] = 'N/A'

			result.append(temp_result)
			process_count+=1	
			percent = ShowProgress(process_count, total_process, process_ratio, current_ratio )
			Process_Queue.put(percent)		
		
		current_ratio += process_ratio
		percent = ShowProgress(process_count, total_process, process_ratio, current_ratio )
		Process_Queue.put(percent)		

		Function_Export_Image_Compare_Test_Result(result, _unique_image_dir,result_file)
		percent = ShowProgress(1, 1)
		Process_Queue.put(percent)
		
	elif scan_type == 'Image and Text':	
		_all_image_dir = _output_dir + '/all_images'
		_unique_image_dir = _output_dir + '/unique_images'
		result_file = _output_dir + '/' + 'Full_Gacha_Test_Result' + '.xlsx'
		
		initFolder(_all_image_dir)
		initFolder(_unique_image_dir)

		_db_dir = os.path.dirname(db_path)

		DB = Function_Import_DB(db_path)
		word_db_list = []
		for element in DB:
			word = element[tess_language]
			word_db_list.append(word.replace(' ','').lower())
		
		current_ratio = 0
		process_ratio = 0.0
		process_count = 0
		percent = ShowProgress(process_ratio, 100)
		current_ratio+=process_ratio
		Process_Queue.put(percent)

		Status_Queue.put('Crop image')
		process_ratio = 0.1
		image_info = Function_Crop_All_Component_And_Text(Process_Queue, image_files, scan_areas, ratio, _all_image_dir, process_ratio, current_ratio)
		current_ratio+=process_ratio

		Status_Queue.put('Filter unique images ('+ str(image_info['count']) + ')')
		#Define weith of the current step
		process_ratio =0.3
		_draft_result = Function_Filter_Unique_DB(Process_Queue, image_info, _unique_image_dir, process_ratio, current_ratio)
		print('_draft_result', _draft_result)

		current_ratio+=process_ratio
		total_process = len(_draft_result)
		percent = ShowProgress(current_ratio, 100)
		Process_Queue.put(percent)		
		result = []
		# Search for Image in DB:
		Status_Queue.put('Scan component images with DB')
		print('Step 1: ', len(_draft_result), len(result))
		#Define weith of the current step
		process_ratio =0.3
		

		for unique_image in _draft_result[:]:
			_match_type = 'none'
			temp_result = {}
			_temp_template = _output_dir + '/' + unique_image['component']
			for row in DB:
				_template = _db_dir + '\\' + row['path']
				if os.path.isfile(_template):
					_image_compare_result = Function_Compare_2_Image(_temp_template, _template)
					if _image_compare_result == True:
						_name = row[tess_language]
						temp_result['image'] = _temp_template
						temp_result['count'] = unique_image['count']
						temp_result['text_raw'] = unique_image['text_raw'][0]
						temp_result['name'] = _name
						temp_result['match_type'] = 'Image compare: Image exact math'
						_draft_result.remove(unique_image)
						result.append(temp_result)
						#break
		
			process_count+=1	
			percent = ShowProgress(process_count, total_process, process_ratio, current_ratio )
			Process_Queue.put(percent)

		#Update progress bar
		current_ratio += process_ratio
		percent = ShowProgress(process_count, total_process, process_ratio, current_ratio )
		Process_Queue.put(percent)		

		print('Step 2: ', len(_draft_result), len(result))

		#Define weith of the current step
		Status_Queue.put('Scan component text by OCR')
		process_ratio = 0.3
		
		error_count = 0
		for unique_image in _draft_result[:]:

			_temp_template = _output_dir + '/' +  unique_image['component']
			key = str(Function_Get_Text_from_Image(tess_path, tess_language, advanced_tessdata_dir_config, _temp_template))
			
			if len(word_db_list)> 0:
				_temp_text = key.replace(' ','').lower()
				if len(_temp_text) == 0:
					_match_type = 'None'
				elif _temp_text in word_db_list:
					# exact match
					_match_type = 'OCR: Text exact'
					print('_key', key)
				else:
					# similarity check
					_dist = len(_temp_text)
					_ratio = 0
					_word = ''
					
					for word in word_db_list:
						Distance = lev_distance.distance(_temp_text, word)		
						Ratio = lev_ratio.normalized_similarity(_temp_text, word)
						if Distance <= _dist and Ratio >= _ratio:
							_dist = Distance
							_ratio = Ratio
							_word = word

					if _dist/len(_temp_text) <= 0.2 and _ratio >= 0.8:
						_index = word_db_list.index(_word)
						_key = word_db_list[_index]
						Status_Queue.put('Text has been corrected from: ' + key + ' to ' + _key)
						key = _key
						_match_type = 'OCR: Text corrected with DB'
					elif _dist/len(_temp_text) < 0.34 and _ratio > 0.66 and len(_temp_text) == len(_word):
						_index = word_db_list.index(_word)
						_key = word_db_list[_index]
						Status_Queue.put('Text has been corrected from: ' + key + ' to ' + _key)
						key = _key
						_match_type = 'OCR: Text corrected with DB'
					else:
						_match_type = 'None'
					
			
			if len(key.replace(' ','')) == 0:
				error_count+=1
				key = '[Error_'+ str(error_count) + ']'
	
			temp_result = {}
			temp_result['image'] = _temp_template
			temp_result['count'] = unique_image['count']
			temp_result['text_raw'] = unique_image['text_raw'][0]
			temp_result['name'] = key
			temp_result['match_type'] = _match_type
	
			result.append(temp_result)
			_draft_result.remove(unique_image)
			process_count+=1	
			percent = ShowProgress(process_count, total_process, process_ratio, current_ratio )
			Process_Queue.put(percent)

		print('Step 3: ', len(_draft_result), len(result))

		Status_Queue.put('Export test result')
		row_height = scan_areas[0][7]
		cell_width = scan_areas[0][2]* (1/6)

		Function_Export_Gacha_Full_Test_Result(result, _unique_image_dir,result_file, cell_width, row_height)
		percent = ShowProgress(1, 1)
		Process_Queue.put(percent)

	elif scan_type == 'DB Create':	
		
		_db_dir = os.path.dirname(db_path)
		_template_dir = _db_dir + '\\all_images'
		_unique_image_dir = _db_dir + '\\template_storage'
	
		DB = Function_Import_DB(db_path)

		initFolder(_template_dir)
		initFolder(_unique_image_dir)

		current_ratio = 0
		process_ratio = 0.0
		process_count = 0
		percent = ShowProgress(process_ratio, 100)
		current_ratio+=process_ratio


		Process_Queue.put(percent)

		Status_Queue.put('Crop image')
		process_ratio = 0.04
		image_info = Function_Crop_All_Component_And_Text(Process_Queue, image_files, scan_areas, ratio, _template_dir, process_ratio, current_ratio)
		
		current_ratio+=process_ratio
		percent = ShowProgress(process_ratio, 100)
		current_ratio+=process_ratio
		Process_Queue.put(percent)

		Status_Queue.put('Filter unique images ('+ str(image_info['count']) + ')')
		process_ratio =0.5

		_draft_result = Function_Filter_Unique_DB(Process_Queue, image_info, _unique_image_dir, process_ratio, current_ratio)

		current_ratio+=process_ratio
		total_process = len(_draft_result)
		
		for component_detail in _draft_result:
	
			key = str(Function_Get_Text_from_Image(tess_path, tess_language, advanced_tessdata_dir_config, component_detail['text_raw'][0]))
			key = key.replace(' ','')
			component_detail['text'] = key
			
			process_count+=1	
			percent = ShowProgress(process_count, total_process, process_ratio, current_ratio )
			Process_Queue.put(percent)		
		rmtree(_template_dir)
		#rmtree(_output_dir)
		current_ratio += process_ratio
		
		Status_Queue.put('Export test result')
		Function_Export_Auto_DB(Status_Queue, _draft_result, tess_language, db_path)
		Status_Queue.put('Append ' + str(len(_draft_result)) + ' row(s) to DB.')
		percent = ShowProgress(1, 1)
		Process_Queue.put(percent)
		

	else:
		Status_Queue.put('Unsupport type')	

def Get_Text_From_Single_Image(tess_path, tess_language, advanced_tessdata_dir_config, input_image, ratio, scan_areas, result_file,):

	pytesseract.pytesseract.tesseract_cmd = tess_path
	_img = Load_Image_by_Ratio(input_image, ratio)
	_result = []
	_output_dir = os.path.dirname(result_file)
	baseName = os.path.basename(input_image)
	sourcename, ext = os.path.splitext(baseName)
	_area_count = 0
	for area in scan_areas:
		_area_count +=1
		imCrop = _img[int(area[1]):int(area[1]+area[3]), int(area[0]):int(area[0]+area[2])]
		_name = _output_dir + '\\' + sourcename + '_' + str(_area_count) + ext
		cv2.imwrite(_name, imCrop)
		
		imCrop = Function_Pre_Processing_Image(imCrop)
		#_name = _output_dir + '\\' + sourcename + '_' + str(_area_count) + '.jpg'
		#cv2.imwrite(_name, imCrop)
		ocr = Get_Text(imCrop, tess_language, advanced_tessdata_dir_config)
		_result.append(ocr)

	baseName = os.path.basename(input_image)
	file_name = os.path.splitext(baseName)[0]
	while True:
		try:
			with open(result_file, 'a', newline='', encoding='utf-8-sig') as csvfile:
				writer = csv.writer(csvfile)
				writer.writerow([file_name] + _result)
				break
		except PermissionError:
			continue

def Function_Get_Text_from_Image(tess_path, tess_language, advanced_tessdata_dir_config, input_image):
	pytesseract.pytesseract.tesseract_cmd = tess_path
	imCrop = cv2.imread(input_image)
	ocr = Get_Text(imCrop, tess_language, advanced_tessdata_dir_config)
	return ocr

def Function_Compare_2_Image(source_image_path, target_image_path):
	if not os.path.isfile(source_image_path):
		print('File not existed: ', source_image_path)
		return False
	if not os.path.isfile(target_image_path):
		print('File not existed: ', target_image_path)
		return False
		
	source_image = cv2.imread(source_image_path)
	source_image = cv2.cvtColor(source_image, cv2.COLOR_BGR2GRAY)	

	target_image = cv2.imread(target_image_path)
	target_image = cv2.cvtColor(target_image, cv2.COLOR_BGR2GRAY)	
	
	result = cv2.matchTemplate(source_image, target_image, cv2.TM_CCOEFF_NORMED)
	(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)
	#print('maxVal', maxVal)
	if maxVal > 0.8:
		return True
	else:
		return False

def Function_Compare_2_Component(source_image_path, target_image_path):
	source_image = cv2.imread(source_image_path)
	source_image = cv2.cvtColor(source_image, cv2.COLOR_BGR2GRAY)
	
	target_image = cv2.imread(target_image_path)
	target_image = cv2.cvtColor(target_image, cv2.COLOR_BGR2GRAY)	
	
	result = cv2.matchTemplate(source_image, target_image, cv2.TM_CCOEFF_NORMED)
	(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)
	#print('maxVal', maxVal)

	if maxVal > 0.5:
		return True
	else:
		return False

def Function_Crop_All_Text(Process_Queue, source_images, scan_areas, ratio, output_dir, start_percent, process_ratio):
	
	total_task = len(scan_areas) * len(source_images)
	amount = 0
	_total_w = 0
	_total_h = 0

	for area in scan_areas:
		_total_w += area[2]
		_total_h += area[3]
	_avg_w = _total_w/(len(scan_areas))
	_avg_h = _total_h/(len(scan_areas))

	info = {}

	for image in source_images:
		_area_count = 0
		baseName = os.path.basename(image)
		sourcename, ext = os.path.splitext(baseName)
		_img = Load_Image_by_Ratio(image, ratio)
		info[sourcename] = []
		for area in scan_areas:
			_area_count +=1
			imCrop = _img[int(area[1]):int(area[1]+_avg_h), int(area[0]):int(area[0]+_avg_w)]
			_name = output_dir + '\\' + sourcename + '_' + str(_area_count) + ext
			
			cv2.imwrite(_name, imCrop)

			image_info = {}
			image_info['link'] = _name
			image_info['area'] = _area_count

			info[sourcename].append(image_info)
			amount+=1

		percent = ShowProgress(amount, total_task, process_ratio, start_percent)
		Process_Queue.put(percent)		

	info['count'] = amount

	return info

def Function_Crop_All_Component_And_Text(Process_Queue, source_images, scan_areas, ratio, output_dir, start_percent, process_ratio):
	
	total_task = len(scan_areas) * len(source_images)
	amount = 0
	_total_w_1 = 0
	_total_h_1 = 0

	_total_w_2 = 0
	_total_h_2 = 0

	for area in scan_areas:
		_total_w_1 += area[2]
		_total_h_1 += area[3]
		_total_w_2 += area[6]
		_total_h_2 += area[7]

	_avg_w_1 = _total_w_1/(len(scan_areas))
	_avg_h_1 = _total_h_1/(len(scan_areas))

	_avg_w_2 = _total_w_2/(len(scan_areas))
	_avg_h_2 = _total_h_2/(len(scan_areas))

	info = {}

	for image in source_images:
		_area_count = 0
		baseName = os.path.basename(image)
		sourcename, ext = os.path.splitext(baseName)
		_img = Load_Image_by_Ratio(image, ratio)
		info[sourcename] = []
		for area in scan_areas:
			_area_count +=1
			# Components:
			imCrop = _img[int(area[5]):int(area[5]+_avg_h_2), int(area[4]):int(area[4]+_avg_w_2)]
			_component_name = output_dir + '\\' + sourcename + '_component_' + str(_area_count) + ext
			cv2.imwrite(_component_name, imCrop)
			# Text:
			imCrop = _img[int(area[1]):int(area[1]+_avg_h_1), int(area[0]):int(area[0]+_avg_w_1)]
			_text_name = output_dir + '\\' + sourcename + '_text_' + str(_area_count) + ext
			cv2.imwrite(_text_name, imCrop)

			image_info = {}
			image_info['component_path'] = _component_name
			image_info['text_path'] = _text_name
			image_info['area'] = _area_count

			info[sourcename].append(image_info)
			amount+=1

		percent = ShowProgress(amount, total_task, process_ratio, start_percent)
		Process_Queue.put(percent)		

	info['count'] = amount

	return info

def Function_Crop_All_Image(Process_Queue, source_images, scan_areas, ratio, output_dir, start_percent, process_ratio):
	
	total_task = len(scan_areas) * len(source_images)
	amount = 0
	_total_w = 0
	_total_h = 0

	for area in scan_areas:
		_total_w += area[6]
		_total_h += area[7]
	_avg_w = _total_w/(len(scan_areas))
	_avg_h = _total_h/(len(scan_areas))

	info = {}

	for image in source_images:
		_area_count = 0
		baseName = os.path.basename(image)
		sourcename, ext = os.path.splitext(baseName)
		_img = Load_Image_by_Ratio(image, ratio)
		info[sourcename] = []
		for area in scan_areas:
			_area_count +=1
			imCrop = _img[int(area[5]):int(area[5]+_avg_h), int(area[4]):int(area[4]+_avg_w)]
			_name = output_dir + '\\' + sourcename + '_' + str(_area_count) + ext
			
			cv2.imwrite(_name, imCrop)

			image_info = {}
			image_info['link'] = _name
			image_info['area'] = _area_count

			info[sourcename].append(image_info)
			amount+=1

		percent = ShowProgress(amount, total_task, process_ratio, start_percent)
		Process_Queue.put(percent)		

	info['count'] = amount

	return info

def Function_Filter_Unique_DB(Process_Queue, all_image_info, unique_images_dir, start_percent, process_ratio):

	#all_image_info['area'] = _area_count	
	unique_images_folder = os.path.basename(unique_images_dir)
	all_component_images = []
	all_text_images = []
	for full_image in all_image_info:
		if full_image != 'count':
			for image in all_image_info[full_image]:
				all_component_images.append(image['component_path'])
				all_text_images.append(image['text_path'])
	
	unique = []
	unique_data = []
	
	process = 0
	all_process = len(all_component_images)

	for component_image in all_component_images:
		baseName = os.path.basename(component_image)
		element = {}
		if len(unique_data) == 0:
			
			element['component'] = unique_images_folder + '\\' + baseName
			_index = all_component_images.index(component_image)
			element['text_raw'] = [all_text_images[_index]]
			element['count'] = 1
			unique_data.append(element)
			unique.append(component_image)
			Export_Unique_Image(component_image, unique_images_dir)
			Export_Unique_Image(all_text_images[_index], unique_images_dir)
		else:
			all_result = False
			for target_image in unique:
				result = Function_Compare_2_Image(component_image, target_image)
				if result == True:
					all_result = True
					for element in unique_data:
						if os.path.basename(element['component']) == os.path.basename(target_image):
							element['count'] +=1
					break
			if all_result == False:
				element['component'] = unique_images_folder + '\\' + baseName
				_index = all_component_images.index(component_image)
	
				element['text_raw'] = [all_text_images[_index]]
				element['count'] = 1
				unique_data.append(element)
				unique.append(component_image)
				Export_Unique_Image(component_image, unique_images_dir)
				Export_Unique_Image(all_text_images[_index], unique_images_dir)
				
		process+=1		
		percent = ShowProgress(process, all_process, start_percent, process_ratio)
		Process_Queue.put(percent)			

	return unique_data

def Function_Filter_Unique_Image(Process_Queue, all_image_dir, unique_images_dir, start_percent, process_ratio):

	_temp_image_files = os.listdir(all_image_dir)
	
	all_images = []
	for image in _temp_image_files:
		image_path = all_image_dir + '\\' + image
		if os.path.isfile(image_path):
			all_images.append(all_image_dir + '\\' + image)
	
	unique = []
	count = {}
	process = 0
	all_process = len(all_images)
	for source_image in all_images:
		baseName = os.path.basename(source_image)
		if len(unique) == 0:
			count[baseName] = 1
			unique.append(source_image)
			Export_Unique_Image(source_image, unique_images_dir)
		else:
			result = False
			for target_image in unique:
				result = Function_Compare_2_Image(source_image, target_image)
				if result == True:
					base_target = os.path.basename(target_image)
					count[base_target] += 1
					break
			if result == False:
				count[baseName] = 1
				unique.append(source_image)
				Export_Unique_Image(source_image, unique_images_dir)
		process+=1		
		percent = ShowProgress(process, all_process, start_percent, process_ratio)
		Process_Queue.put(percent)			
	return count
	
def Export_Unique_Image(path, new_folder):
	#unique_image = cv2.imread(path)
	baseName = os.path.basename(path)
	new_path = new_folder +'\\' + baseName
	copyfile(path, new_path)
	#cv2.imwrite(new_name, unique_image)

def Function_Import_DB(db_path):
	if not os.path.isfile(db_path):
		return []
	_db_dir = os.path.dirname(db_path)
	_all_db = []
	col_name = ['eng', 'kor', 'path']
	with open(db_path, newline='', encoding='utf-8-sig') as csvfile:
		all_components = csv.DictReader(csvfile)
		for component in all_components:
			_db_entry = {}
			for key_name in component:
				if key_name == 'path':
					_temp_path = _db_dir + '\\' + component['path']
					if os.path.isfile(_temp_path):
						_db_entry['path'] = component['path']
					else:
						_db_entry['path'] = ''
				else:
					_db_entry[key_name] = component[key_name]
			_all_db.append(_db_entry)
	
	return _all_db

def Function_Export_Gacha_Test_Result(result_obj, image_dir, result_path, cell_width, row_height):

	all_match_color = Color(rgb='ADF7B6')
	all_match_fill = PatternFill(patternType='solid', fgColor=all_match_color)
	corrected_color = Color(rgb='A0CED9')
	corrected_fill = PatternFill(patternType='solid', fgColor=corrected_color)
	none_color = Color(rgb='FFEE93')
	none_fill = PatternFill(patternType='solid', fgColor=none_color)

	summary = Workbook()
	ws =  summary.active
	ws.title = 'Summary'
	Header = ['Index', 'Component', 'Amount', 'Image']
	Col = 1
	Row = 2
	for Par in Header:
		ws.cell(row=Row, column=Col).value = Par
		Col +=1
	Row +=1

	ws.cell(row=2, column=6).fill = all_match_fill
	ws.cell(row=2, column=7).value = "Component name found in DB"
	ws.cell(row=3, column=6).fill = corrected_fill
	ws.cell(row=3, column=7).value = "Component name corrected by using DB"
	ws.cell(row=4, column=6).fill = none_fill
	ws.cell(row=4, column=7).value = "Component name not found in DB"
	
	_index = 0
	for component in result_obj:
		_index+=1
		ws.cell(row=Row, column=1).value = _index

		ws.cell(row=Row, column=2).value = component
		_match_type = result_obj[component]['match_type']
		for col in range (1, 4):
			if _match_type == 'exact':
				ws.cell(row=Row, column=2).fill = all_match_fill
			elif _match_type == 'corrected':
				ws.cell(row=Row, column=2).fill = corrected_fill
			else:
				ws.cell(row=Row, column=2).fill = none_fill

		count = result_obj[component]['value']
		ws.cell(row=Row, column=3).value = count
		image = image_dir + '\\' + result_obj[component]['image']
		cell_image = Image(image)
		cell_image.anchor = 'D' + str(Row)
		ws.add_image(cell_image)
		
		if row_height:
			ws.row_dimensions[Row].height = row_height
		if cell_width:
			ws.column_dimensions['D'].width = cell_width
	
		Row +=1

	Tab = Table(displayName="Summary", ref="A2:" + "D" + str(Row-1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	Tab.tableStyleInfo = style
	ws.add_table(Tab)
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))	

	column_letters = ['A', 'B', 'C', 'D']
	for column_letter in column_letters:
		ws.column_dimensions[column_letter].bestFit = True

	summary.save(result_path)	
	summary.close()

def Function_Export_Gacha_Full_Test_Result(result_obj, image_dir, result_path, cell_width, row_height):

	all_match_color = Color(rgb='ADF7B6')
	all_match_fill = PatternFill(patternType='solid', fgColor=all_match_color)
	corrected_color = Color(rgb='A0CED9')
	corrected_fill = PatternFill(patternType='solid', fgColor=corrected_color)
	none_color = Color(rgb='FFEE93')
	none_fill = PatternFill(patternType='solid', fgColor=none_color)

	summary = Workbook()
	ws =  summary.active
	ws.title = 'Summary'
	Header = ['Index', 'Image', 'Amount', 'Name (From DB)', 'Raw text', 'Match Type']
	Col = 1
	Row = 2
	for Par in Header:
		ws.cell(row=Row, column=Col).value = Par
		Col +=1
	Row +=1

	ws.cell(row=2, column=8).fill = all_match_fill
	ws.cell(row=2, column=9).value = "Component name or image found in DB"
	ws.cell(row=3, column=8).fill = corrected_fill
	ws.cell(row=3, column=9).value = "Component name corrected by using DB"
	ws.cell(row=4, column=8).fill = none_fill
	ws.cell(row=4, column=9).value = "Component name not found in DB"
	

	_index = 0

	for component in result_obj:
		_index+=1
		ws.cell(row=Row, column=1).value = _index

		cell_image = Image(component['image'])
		cell_image.anchor = 'B' + str(Row)
		ws.add_image(cell_image)

		count = component['count']
		ws.cell(row=Row, column=3).value = count
	
		if row_height:
			ws.row_dimensions[Row].height = row_height
		if cell_width:
			ws.column_dimensions['B'].width = cell_width
			ws.column_dimensions['E'].width = cell_width
		ws.column_dimensions['F'].width = 32
		ws.cell(row=Row, column=4).value = component['name']

		cell_image = Image(component['text_raw'])
		cell_image.anchor = 'E' + str(Row)
		ws.add_image(cell_image)

		_match_type = component['match_type']
		ws.cell(row=Row, column=6).value = _match_type
		for col in range (1, 7):
			if _match_type in ['Image compare: Image exact math', 'OCR: Text exact']:
				ws.cell(row=Row, column=col).fill = all_match_fill
			elif _match_type == 'OCR: Text corrected with DB':
				ws.cell(row=Row, column=col).fill = corrected_fill
			else:
				ws.cell(row=Row, column=col).fill = none_fill
		Row +=1

	Tab = Table(displayName="Summary", ref="A2:" + "F" + str(Row-1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	Tab.tableStyleInfo = style
	ws.add_table(Tab)
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))	

	column_letters = ['A', 'B', 'C', 'D', 'E', 'F']
	for column_letter in column_letters:
		ws.column_dimensions[column_letter].bestFit = True


	summary.save(result_path)	
	summary.close()

def Function_Export_Image_Compare_Test_Result(result_obj, image_dir, result_path):

	summary = Workbook()
	ws =  summary.active
	ws.title = 'Summary'
	Header = ['Index', 'Image', 'Amount', 'Name (From DB)', 'Match Type']
	ws.cell(row=3, column=5).value = 'Image compare: Image exact math'
	Col = 1
	Row = 2
	for Par in Header:
		ws.cell(row=Row, column=Col).value = Par
		Col +=1
	Row +=1
	
	_index = 0

	for component in result_obj:
		_index+=1
		ws.cell(row=Row, column=1).value = _index

		cell_image = Image(component['image'])
		curent_w = cell_image.width
		curent_h = cell_image.height
		_size = 128
		_ratio = _size/curent_w
		cell_image.width = _size
		cell_image.height = curent_h*_ratio
		cell_image.anchor = 'B' + str(Row)
		ws.add_image(cell_image)

		count = component['count']
		ws.cell(row=Row, column=3).value = count
	
		ws.row_dimensions[Row].height = int(curent_h*_ratio/1.3)
		ws.column_dimensions['B'].width = 19
		if 'name' in component:
			ws.cell(row=Row, column=4).value = component['name']

		#ws.cell(row=Row, column=6).value = component['match_type']
	
		Row +=1

	Tab = Table(displayName="Summary", ref="A2:" + "D" + str(Row-1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	Tab.tableStyleInfo = style
	ws.add_table(Tab)
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))	

	column_letters = ['A', 'B', 'C', 'D']
	for column_letter in column_letters:
		ws.column_dimensions[column_letter].bestFit = True

	# Split path, filename and extension
	_path, _filename = os.path.split(result_path)
	_raw_filename, _extension = os.path.splitext(_filename)
	# Add timestamp to _raw_filename
	_filename = _raw_filename + '_' + timestamp + _extension
	# Join path and _filename
	result_path = os.path.join(_path, _filename)
	summary.save(result_path)	
	summary.close()

def Function_Export_Auto_DB(status_queue, result_obj, tess_language, result_path):
	_exist = os.path.isfile(result_path)
	current_db = Function_Import_DB(result_path)
	_db_dir = os.path.dirname(result_path)
	print('_db_dir', _db_dir)
	with open(result_path, 'a', newline='', encoding='utf-8-sig') as csvfile:
		fieldnames = ['kor', 'eng', 'path']
		writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
		if not _exist:
			writer.writeheader()

		temp_db = []	
		for component_details in result_obj:
			result = False

			for row in current_db:
				_template = _db_dir + '\\' + row['path']
				_new_template = _db_dir + '\\' + component_details['component']
				result = Function_Compare_2_Image(_new_template, _template)
				if result == True:
					#if row[tess_language] == "" and component_details['text'] != '':
					#	print('Updating text to existed DB row.')
					#	status_queue.put('Updating text to existed DB row')
					#	row[tess_language] =component_details['text']
					break
				else:
					result = False
			if result == False:
				for row in temp_db:
					_template = _db_dir + '\\' + row['path']
					_new_template = _db_dir + '\\' + component_details['component']
					result = Function_Compare_2_Image(_new_template, _template)
					if result == True:
						break
					else:
						result = False
		
			if result == False:
				print('Append new row to existed DB.')
				_new_row = {'path': component_details['component'], tess_language: component_details['text']}
				writer.writerow(_new_row)
				temp_db.append(_new_row)
			else:
				print('Entry existed in existed DB.')
				status_queue.put('Entry existed in existed DB.')


def Function_Analyze_Gacha_Data(_raw_data, col_name):
	_output_dir = os.path.dirname(_raw_data)
	analyze_result_file = _output_dir + "/analyze_result.csv"
	_gacha = {}
	
	with open(_raw_data, newline='', encoding='utf-8-sig') as csvfile:
		reader = csv.DictReader(csvfile)
		for row in reader:
			for component in col_name:
				component_name = row[component]
				if component_name in _gacha:
					_gacha[component_name] +=1
				else:
					_gacha[component_name] =1

	with open(analyze_result_file, 'a', newline='', encoding='utf-8-sig') as csvfile:
		fieldnames = ['Components', 'Amount']
		writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
		writer.writeheader()
		for component in _gacha:
			writer.writerow({'Components': component, 'Amount': _gacha[component]})
	print('Analyze done')	



def Get_Text(img, tess_language, tessdata_dir_config):
	ocr = pytesseract.image_to_string(img, lang = tess_language, config=tessdata_dir_config)
	ocr = ocr.replace("\n", "")
	ocr = ocr.replace("\r", "")  
	ocr = ocr.replace("\x0c", "") 
	return ocr

###########################################################################################
# Process function - Preview scan
###########################################################################################

def Function_Preview_Scan(
	Result_Queue, Process_Queue, tess_path, tess_language, test_data, image_files, ratio, scan_areas, **kwargs):
	pytesseract.pytesseract.tesseract_cmd = tess_path
	tessdata_dir_config = '--tessdata-dir ' + '"' + test_data + '"'
	if tess_language == '':
		tess_language = 'kor'
	
	_result = []
	_counter = 0
	_total = len(scan_areas)
	_result = []

	_img = Load_Image_by_Ratio(image_files, ratio)


	for area in scan_areas:
		imCrop = _img[int(area[1]):int(area[1]+area[3]), int(area[0]):int(area[0]+area[2])]
		imCrop = Function_Pre_Processing_Image(imCrop)
		
		ocr = Get_Text(imCrop, tess_language, tessdata_dir_config)
	
		_result.append(ocr)
		_counter+=1	
		Process_Queue.put(int(_counter*1000/_total))

	Result_Queue.put(_result)
	
def Load_Image_by_Ratio(image_path, resolution):
	
	_img = cv2.imread(image_path)
	
	(_h, _w) = _img.shape[:2]
	
	_ratio = resolution / _h
	if _ratio != 1:
		width = int(_img.shape[1] * _ratio)
		height = int(_img.shape[0] * _ratio)
		dim = (width, height)
		_img = cv2.resize(_img, dim, interpolation = cv2.INTER_AREA)
	
	return _img



def Function_Pre_Processing_Image(img):
	img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
	#img = cv2.resize(img, None, fx=0.5, fy=0.5, interpolation=cv2.INTER_AREA)
	img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
	img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_LINEAR)
	img = cv2.blur(img,(5,5))
	img = cv2.GaussianBlur(img, (5, 5), 0)
	img = cv2.medianBlur(img, 3)
	img = cv2.bilateralFilter(img,9,75,75)
	#cv2.threshold(img,127,255,cv2.THRESH_BINARY)
	img = image_smoothening(img)

	return	img

def image_smoothening(img):
	BINARY_THREHOLD = 100
	ret1, th1 = cv2.threshold(img, BINARY_THREHOLD, 255, cv2.THRESH_BINARY)
	ret2, th2 = cv2.threshold(th1, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
	#blur = cv2.GaussianBlur(th2, (1, 1), 0)
	ret3, th3 = cv2.threshold(th2, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
	return th3


def ShowProgress(Counter, TotalProcess, share=1, start_value=0):
	#print(locals())
	percent = int(1000 * Counter * share/ TotalProcess) + int(start_value*1000)
	#print("Current progress: " +  str(Counter) + '/ ' + str(TotalProcess))
	#print('Percent:', percent)
	return percent

def Function_Get_TimeStamp():
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))
	return timestamp

def initFolder(dir_path):
	'''
	Create the config folder incase it's not existed
	'''
	if not os.path.isdir(dir_path):
		try:
			os.mkdir(dir_path)
		except OSError:
			print ("Creation of the directory %s failed" % dir_path)
		else:
			print ("Successfully created the directory %s " % dir_path)
		#Check local database

###########################################################################################
# Main loop
###########################################################################################



def main():
	global WIDTH, HEIGHT
	Process_Queue = Queue()
	Result_Queue = Queue()
	Status_Queue = Queue()
	Debug_Queue = Queue()
	
	MyManager = Manager()
	Default_Manager = MyManager.list()
	
	root = Tk()
	WIDTH = root.winfo_screenwidth()
	HEIGHT = root.winfo_screenheight()
	
	My_Queue = {}
	My_Queue['Process_Queue'] = Process_Queue
	My_Queue['Result_Queue'] = Result_Queue
	My_Queue['Status_Queue'] = Status_Queue
	My_Queue['Debug_Queue'] = Debug_Queue
	
	My_Manager = {}
	My_Manager['Default_Manager'] = Default_Manager

	OCR_Project(root, Queue = My_Queue, Manager = My_Manager,)
	
	#root.overrideredirect(1)
	root.attributes("-alpha", 0.98)

	root.mainloop()  


if __name__ == '__main__':
	if sys.platform.startswith('win'):
		multiprocessing.freeze_support()

	main()
